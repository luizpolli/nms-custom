"""Excel report generator using openpyxl.

Produces multi-sheet XLSX workbooks from live database data.
All public methods are async and return raw bytes ready for HTTP streaming.
"""

from __future__ import annotations

import io
from collections import defaultdict
from collections.abc import Callable, Sequence
from datetime import UTC, datetime

import openpyxl
from loguru import logger
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.alarm import Alarm
from app.models.device import Device
from app.models.inventory import Inventory
from app.models.ios_version import IOSVersion
from app.models.kpi import KPI
from app.models.kpi_threshold import KPIThreshold
from app.models.monitoring_policy import MonitoringPolicy
from app.models.service import Service, ServiceScoreSnapshot
from app.services.reports.consolidation import (
    BUCKET_SECONDS,
    BucketSize,
    ConsolidationFn,
    bucketize,
    percentile,
    summary,
)

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def _style_header_row(ws: Worksheet) -> None:
    """Bold + blue header row, freeze top row."""
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def _autosize(ws: Worksheet) -> None:
    """Set column width to max content length (capped at 60)."""
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = max((len(str(c.value or "")) for c in col_cells), default=8)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)


def _finalise(ws: Worksheet) -> None:
    _style_header_row(ws)
    _autosize(ws)


def _wb_to_bytes(wb: openpyxl.Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class ExcelReporter:
    """Generates XLSX reports from the NMS database."""

    def __init__(self, session_factory: Callable[[], AsyncSession]) -> None:
        self._sf = session_factory

    # ------------------------------------------------------------------
    # Device inventory report
    # ------------------------------------------------------------------

    async def device_inventory_report(self) -> bytes:
        """Three-sheet workbook: Devices, Inventory, IOS Versions."""
        async with self._sf() as session:
            devices = (await session.execute(select(Device))).scalars().all()
            inventories = (await session.execute(select(Inventory))).scalars().all()
            ios_versions = (await session.execute(select(IOSVersion))).scalars().all()

        logger.info("device_inventory_report: {} devices", len(devices))
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # type: ignore[arg-type]

        self._sheet_devices(wb, devices)
        self._sheet_inventory(wb, inventories, {d.id: d.name for d in devices})
        self._sheet_ios_versions(wb, ios_versions, {d.id: d.name for d in devices})

        return _wb_to_bytes(wb)

    def _sheet_devices(self, wb: openpyxl.Workbook, devices: Sequence[Device]) -> None:
        ws = wb.create_sheet("Devices")
        ws.append(["ID", "Name", "IP Address", "Vendor", "Model", "OS Type", "Status", "Location"])
        for d in devices:
            ws.append([str(d.id), d.name, d.ip_address, d.vendor, d.model, d.os_type, d.status, d.location])
        _finalise(ws)

    def _sheet_inventory(
        self, wb: openpyxl.Workbook, rows: Sequence[Inventory], device_names: dict
    ) -> None:
        ws = wb.create_sheet("Inventory")
        ws.append(["Device", "Serial", "HW Model", "FW Version", "Port Count", "Uptime (s)", "Mem Total", "Mem Free"])
        for inv in rows:
            ws.append([
                device_names.get(inv.device_id, str(inv.device_id)),
                inv.serial_number, inv.hardware_model, inv.firmware_version,
                inv.port_count, inv.uptime_seconds, inv.memory_total, inv.memory_free,
            ])
        _finalise(ws)

    def _sheet_ios_versions(
        self, wb: openpyxl.Workbook, rows: Sequence[IOSVersion], device_names: dict
    ) -> None:
        ws = wb.create_sheet("IOS Versions")
        ws.append(["Device", "Version", "Image File", "Platform", "Is EOL", "Is EOS"])
        for iv in rows:
            ws.append([
                device_names.get(iv.device_id, str(iv.device_id)),
                iv.version, iv.image_file, iv.platform, iv.is_eol, iv.is_eos,
            ])
        _finalise(ws)

    # ------------------------------------------------------------------
    # KPI report
    # ------------------------------------------------------------------

    async def kpi_report(
        self,
        since: datetime,
        until: datetime,
        device_ids: list | None = None,
        bucket: BucketSize = "raw",
        consolidation: ConsolidationFn = "avg",
    ) -> bytes:
        """One sheet per kpi_type; columns = devices, rows = bucketed timestamps.

        ``bucket`` selects a Cricket-style consolidation period (5min/15min/1h/1d) and
        ``consolidation`` selects the per-bucket function (avg/min/max/p95/p99).
        """
        async with self._sf() as session:
            q = select(KPI).where(KPI.timestamp >= since, KPI.timestamp <= until)
            if device_ids:
                q = q.where(KPI.device_id.in_(device_ids))
            kpis = (await session.execute(q)).scalars().all()
            devices = (await session.execute(select(Device))).scalars().all()

        device_names = {d.id: d.name for d in devices}
        logger.info(
            "kpi_report: {} rows in [{}, {}] bucket={} cf={}",
            len(kpis), since, until, bucket, consolidation,
        )

        # Group raw samples by kpi_type → device → list of (timestamp, value)
        raw: dict[str, dict[str, list[tuple[datetime, float]]]] = defaultdict(lambda: defaultdict(list))
        for kpi in kpis:
            dev = device_names.get(kpi.device_id, str(kpi.device_id))
            raw[kpi.kpi_type][dev].append((kpi.timestamp, kpi.value))

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # type: ignore[arg-type]

        meta_ws = wb.create_sheet("Settings")
        meta_ws.append(["Field", "Value"])
        meta_ws.append(["Period", f"{since.isoformat()} → {until.isoformat()}"])
        meta_ws.append(["Bucket", bucket])
        meta_ws.append(["Consolidation", consolidation])
        meta_ws.append(["Devices", len(device_ids) if device_ids else "all"])
        _finalise(meta_ws)

        for kpi_type, dev_map in sorted(raw.items()):
            ws = wb.create_sheet(kpi_type[:31])
            consolidated: dict[str, dict[datetime, float]] = {}
            for dev, samples in dev_map.items():
                rolled = bucketize(samples, bucket, consolidation)
                consolidated[dev] = {ts: value for ts, value in rolled}
            all_devs = sorted(consolidated.keys())
            all_ts = sorted({ts for d in consolidated.values() for ts in d})
            ws.append(["Timestamp"] + all_devs)
            for ts in all_ts:
                ws.append([ts.isoformat()] + [consolidated[dev].get(ts) for dev in all_devs])
            _finalise(ws)

            stats_ws = wb.create_sheet(f"{kpi_type[:25]} stats")
            stats_ws.append(["Device", "Samples", "Avg", "Min", "Max", "P95", "P99"])
            for dev in all_devs:
                values = [v for _, v in dev_map[dev]]
                s = summary(values)
                stats_ws.append([
                    dev, s["samples"], s["avg"], s["min"], s["max"], s["p95"], s["p99"],
                ])
            _finalise(stats_ws)

        if len(wb.sheetnames) == 1:
            wb.create_sheet("No Data")

        return _wb_to_bytes(wb)

    # ------------------------------------------------------------------
    # KPI Top-N report (PPM-style)
    # ------------------------------------------------------------------

    async def kpi_top_n_report(
        self,
        since: datetime,
        until: datetime,
        top_n: int = 10,
        consolidation: ConsolidationFn = "p95",
        kpi_types: list[str] | None = None,
    ) -> bytes:
        """One sheet per KPI listing the Top-N devices by chosen consolidation."""
        async with self._sf() as session:
            q = select(KPI).where(KPI.timestamp >= since, KPI.timestamp <= until)
            if kpi_types:
                q = q.where(KPI.kpi_type.in_(kpi_types))
            kpis = (await session.execute(q)).scalars().all()
            devices = (await session.execute(select(Device))).scalars().all()

        device_names = {d.id: d.name for d in devices}
        grouped: dict[str, dict[str, list[float]]] = defaultdict(lambda: defaultdict(list))
        for kpi in kpis:
            dev = device_names.get(kpi.device_id, str(kpi.device_id))
            grouped[kpi.kpi_type][dev].append(kpi.value)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # type: ignore[arg-type]

        if not grouped:
            wb.create_sheet("No Data")
            return _wb_to_bytes(wb)

        for kpi_type, dev_map in sorted(grouped.items()):
            ws = wb.create_sheet(f"Top {kpi_type[:25]}")
            ws.append(["Rank", "Device", "Samples", "Avg", "Min", "Max", "P95", "P99"])
            ranked = sorted(
                dev_map.items(),
                key=lambda item: (_pick(item[1], consolidation) or float("-inf")),
                reverse=True,
            )[:top_n]
            for idx, (dev, values) in enumerate(ranked, start=1):
                s = summary(values)
                ws.append([
                    idx, dev, s["samples"], s["avg"], s["min"], s["max"], s["p95"], s["p99"],
                ])
            _finalise(ws)
        return _wb_to_bytes(wb)

    # ------------------------------------------------------------------
    # KPI trends report (Cricket-style: hourly/daily/weekly summary side-by-side)
    # ------------------------------------------------------------------

    async def kpi_trends_report(
        self,
        since: datetime,
        until: datetime,
        device_ids: list | None = None,
        buckets: list[BucketSize] | None = None,
    ) -> bytes:
        """One sheet per device listing avg/min/max/p95 per bucket per KPI."""
        bucket_list: list[BucketSize] = buckets or ["1h", "1d", "1w"]
        for b in bucket_list:
            if b not in BUCKET_SECONDS:
                raise ValueError(f"Unsupported bucket: {b!r}")

        async with self._sf() as session:
            q = select(KPI).where(KPI.timestamp >= since, KPI.timestamp <= until)
            if device_ids:
                q = q.where(KPI.device_id.in_(device_ids))
            kpis = (await session.execute(q)).scalars().all()
            devices = (await session.execute(select(Device))).scalars().all()

        device_names = {d.id: d.name for d in devices}
        per_device: dict[str, dict[str, list[tuple[datetime, float]]]] = defaultdict(lambda: defaultdict(list))
        for kpi in kpis:
            dev = device_names.get(kpi.device_id, str(kpi.device_id))
            per_device[dev][kpi.kpi_type].append((kpi.timestamp, kpi.value))

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # type: ignore[arg-type]

        if not per_device:
            wb.create_sheet("No Data")
            return _wb_to_bytes(wb)

        for dev in sorted(per_device):
            ws = wb.create_sheet(dev[:31])
            ws.append(["KPI", "Bucket", "Periods", "Avg", "Min", "Max", "P95"])
            for kpi_type in sorted(per_device[dev]):
                samples = per_device[dev][kpi_type]
                for bucket in bucket_list:
                    rolled = bucketize(samples, bucket, "avg")
                    if not rolled:
                        continue
                    values = [v for _, v in rolled]
                    s = summary(values)
                    ws.append([
                        kpi_type, bucket, s["samples"],
                        s["avg"], s["min"], s["max"], s["p95"],
                    ])
            _finalise(ws)
        return _wb_to_bytes(wb)

    # ------------------------------------------------------------------
    # Baseline comparison (PPM-style)
    # ------------------------------------------------------------------

    async def baseline_comparison_report(
        self,
        since: datetime,
        until: datetime,
        baseline_periods: int = 4,
        device_ids: list | None = None,
    ) -> bytes:
        """Compare the current period KPI averages against ``baseline_periods`` prior windows."""
        if baseline_periods < 1:
            baseline_periods = 1
        windows = [(since, until)]
        period = until - since
        for i in range(1, baseline_periods + 1):
            windows.append((since - period * i, until - period * i))
        earliest = min(w[0] for w in windows)
        latest = max(w[1] for w in windows)

        async with self._sf() as session:
            q = select(KPI).where(KPI.timestamp >= earliest, KPI.timestamp <= latest)
            if device_ids:
                q = q.where(KPI.device_id.in_(device_ids))
            kpis = (await session.execute(q)).scalars().all()
            devices = (await session.execute(select(Device))).scalars().all()

        device_names = {d.id: d.name for d in devices}
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Baseline"
        header = ["Device", "KPI", "Stat"] + [f"P{i}" for i in range(len(windows))]
        header += ["Delta vs P1", "Delta %"]
        ws.append(header)
        ws.append(["", "", "", *(f"{a.date()}..{b.date()}" for a, b in windows), "", ""])

        bucketed: dict[tuple[str, str, int], list[float]] = defaultdict(list)
        for kpi in kpis:
            dev = device_names.get(kpi.device_id, str(kpi.device_id))
            ts = kpi.timestamp
            for idx, (a, b) in enumerate(windows):
                a_aware = a if a.tzinfo else a.replace(tzinfo=UTC)
                b_aware = b if b.tzinfo else b.replace(tzinfo=UTC)
                ts_aware = ts if ts.tzinfo else ts.replace(tzinfo=UTC)
                if a_aware <= ts_aware <= b_aware:
                    bucketed[(dev, kpi.kpi_type, idx)].append(kpi.value)
                    break

        pairs = {(dev, kpi) for (dev, kpi, _) in bucketed}
        for dev, kpi in sorted(pairs):
            for stat in ("avg", "p95"):
                row: list = [dev, kpi, stat]
                values_per_period: list[float | None] = []
                for idx, _ in enumerate(windows):
                    samples = bucketed.get((dev, kpi, idx), [])
                    if not samples:
                        values_per_period.append(None)
                    elif stat == "avg":
                        values_per_period.append(sum(samples) / len(samples))
                    else:
                        values_per_period.append(percentile(samples, 95))
                row.extend(values_per_period)
                current = values_per_period[0]
                baseline = values_per_period[1] if len(values_per_period) > 1 else None
                if current is not None and baseline:
                    delta = current - baseline
                    pct = (delta / baseline * 100.0) if baseline else None
                    row.append(round(delta, 4))
                    row.append(round(pct, 2) if pct is not None else None)
                else:
                    row.extend([None, None])
                ws.append(row)
        _finalise(ws)

        meta = wb.create_sheet("Windows")
        meta.append(["Index", "From", "To", "Label"])
        for idx, (a, b) in enumerate(windows):
            label = "current" if idx == 0 else f"baseline -{idx}"
            meta.append([idx, a.isoformat(), b.isoformat(), label])
        _finalise(meta)
        return _wb_to_bytes(wb)

    # ------------------------------------------------------------------
    # TCA / threshold crossing alerts
    # ------------------------------------------------------------------

    async def tca_report(self, since: datetime, until: datetime) -> bytes:
        """List KPI samples that crossed any configured threshold within the window."""
        async with self._sf() as session:
            thresholds = (await session.execute(select(KPIThreshold))).scalars().all()
            kpis = (
                await session.execute(
                    select(KPI).where(KPI.timestamp >= since, KPI.timestamp <= until)
                )
            ).scalars().all()
            devices = (await session.execute(select(Device))).scalars().all()

        device_names = {d.id: d.name for d in devices}
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Threshold Definitions"
        ws.append([
            "Threshold", "KPI", "Operator", "Value", "Clear Value",
            "Severity", "Auto Clear", "Enabled", "Description",
        ])
        for t in thresholds:
            ws.append([
                t.name, t.kpi_type, t.operator, t.value, t.clear_value,
                t.severity, t.auto_clear, t.enabled, t.description,
            ])
        _finalise(ws)

        cross_ws = wb.create_sheet("Crossings")
        cross_ws.append([
            "Timestamp", "Device", "KPI", "Value", "Threshold", "Operator",
            "Severity", "Threshold Name",
        ])
        thresholds_by_type: dict[str, list[KPIThreshold]] = defaultdict(list)
        for t in thresholds:
            if t.enabled:
                thresholds_by_type[t.kpi_type].append(t)

        rows = 0
        for kpi in kpis:
            for t in thresholds_by_type.get(kpi.kpi_type, []):
                if _operator_match(t.operator, kpi.value, t.value):
                    cross_ws.append([
                        kpi.timestamp.isoformat() if kpi.timestamp else "",
                        device_names.get(kpi.device_id, str(kpi.device_id)),
                        kpi.kpi_type,
                        kpi.value,
                        t.value,
                        t.operator,
                        t.severity,
                        t.name,
                    ])
                    rows += 1
        if rows == 0:
            cross_ws.append(["—", "—", "—", "—", "—", "—", "—", "No crossings in period"])
        _finalise(cross_ws)
        return _wb_to_bytes(wb)

    # ------------------------------------------------------------------
    # Alarm report
    # ------------------------------------------------------------------

    async def alarm_report(self, since: datetime, until: datetime) -> bytes:
        """Two sheets: Alarms detail and Summary by severity."""
        async with self._sf() as session:
            q = select(Alarm).where(Alarm.first_seen >= since, Alarm.last_seen <= until)
            alarms = (await session.execute(q)).scalars().all()

        logger.info("alarm_report: {} alarms in [{}, {}]", len(alarms), since, until)
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # type: ignore[arg-type]

        self._sheet_alarms(wb, alarms)
        self._sheet_alarm_summary(wb, alarms)
        return _wb_to_bytes(wb)

    def _sheet_alarms(self, wb: openpyxl.Workbook, alarms: Sequence[Alarm]) -> None:
        ws = wb.create_sheet("Alarms")
        ws.append([
            "ID", "Source Host", "Severity", "Category", "State",
            "Event Type", "Message", "First Seen", "Last Seen", "Occurrences",
        ])
        for a in alarms:
            ws.append([
                str(a.id), a.source_host, a.severity, a.category, a.state,
                a.event_type, a.message,
                a.first_seen.isoformat() if a.first_seen else None,
                a.last_seen.isoformat() if a.last_seen else None,
                a.occurrence_count,
            ])
        _finalise(ws)

    def _sheet_alarm_summary(self, wb: openpyxl.Workbook, alarms: Sequence[Alarm]) -> None:
        ws = wb.create_sheet("Summary")
        counts: dict[str, int] = defaultdict(int)
        for a in alarms:
            counts[a.severity] += 1
        ws.append(["Severity", "Count"])
        for severity, count in sorted(counts.items()):
            ws.append([severity, count])
        _finalise(ws)

    # ------------------------------------------------------------------
    # Monitoring policy report
    # ------------------------------------------------------------------

    async def monitoring_policy_report(self) -> bytes:
        """Monitoring policy configuration and execution status."""
        async with self._sf() as session:
            policies = (await session.execute(select(MonitoringPolicy).order_by(MonitoringPolicy.interval_seconds.asc()))).scalars().all()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Monitoring Policies"
        ws.append([
            "Name", "Type", "Enabled", "Interval", "Target", "Custom OIDs",
            "Last Run", "Next Run", "Last Status", "Last Error", "Description",
        ])
        for policy in policies:
            ws.append([
                policy.name,
                policy.policy_type,
                policy.enabled,
                _format_interval(policy.interval_seconds),
                "All devices" if policy.target_all_devices else f"{len(policy.device_ids)} selected devices",
                len(policy.metric_oids or []),
                policy.last_run_at.isoformat() if policy.last_run_at else None,
                policy.next_run_at.isoformat() if policy.next_run_at else None,
                policy.last_status,
                policy.last_error,
                policy.description,
            ])
        _finalise(ws)
        return _wb_to_bytes(wb)

    # ------------------------------------------------------------------
    # Assurance / service trend reports
    # ------------------------------------------------------------------

    async def assurance_trend_report(
        self,
        since: datetime,
        until: datetime,
        bucket_minutes: int = 15,
    ) -> bytes:
        """Network-wide assurance score trend bucketed over the window.

        Two sheets: Summary (per-bucket avg/min/max + service/sample counts)
        and Per-Service (one row per service per bucket).
        """
        if bucket_minutes < 1:
            bucket_minutes = 1
        since_aware = since if since.tzinfo else since.replace(tzinfo=UTC)
        until_aware = until if until.tzinfo else until.replace(tzinfo=UTC)

        async with self._sf() as session:
            snapshots = (
                await session.execute(
                    select(ServiceScoreSnapshot)
                    .where(ServiceScoreSnapshot.captured_at >= since_aware)
                    .where(ServiceScoreSnapshot.captured_at <= until_aware)
                    .order_by(ServiceScoreSnapshot.captured_at.asc())
                )
            ).scalars().all()
            services = (await session.execute(select(Service))).scalars().all()

        service_names = {s.id: s.name for s in services}
        bucket_secs = bucket_minutes * 60
        since_ts = since_aware.timestamp()

        per_bucket: dict[int, list[tuple]] = defaultdict(list)
        for snap in snapshots:
            ts = snap.captured_at
            ts_aware = ts if ts.tzinfo else ts.replace(tzinfo=UTC)
            offset = ts_aware.timestamp() - since_ts
            key = int(offset // bucket_secs)
            per_bucket[key].append((snap.service_id, snap.score, snap.health_state))

        wb = openpyxl.Workbook()
        summary_ws = wb.active
        summary_ws.title = "Assurance Trend"
        summary_ws.append([
            "Bucket Start", "Avg Score", "Min Score", "Max Score",
            "Sample Count", "Service Count",
        ])
        for key in sorted(per_bucket):
            entries = per_bucket[key]
            scores = [e[1] for e in entries]
            bucket_start = datetime.fromtimestamp(since_ts + key * bucket_secs, tz=UTC)
            summary_ws.append([
                bucket_start.isoformat(),
                round(sum(scores) / len(scores), 2),
                int(min(scores)),
                int(max(scores)),
                len(scores),
                len({e[0] for e in entries}),
            ])
        _finalise(summary_ws)

        detail_ws = wb.create_sheet("Per-Service")
        detail_ws.append([
            "Bucket Start", "Service", "Avg Score", "Min Score", "Max Score",
            "Sample Count", "Worst Health State",
        ])
        _severity_rank = {"healthy": 0, "degraded": 1, "critical": 2, "down": 3}
        for key in sorted(per_bucket):
            bucket_start = datetime.fromtimestamp(since_ts + key * bucket_secs, tz=UTC)
            per_service: dict = defaultdict(list)
            for sid, score, state in per_bucket[key]:
                per_service[sid].append((score, state))
            for sid in per_service:
                samples = per_service[sid]
                scores = [s for s, _ in samples]
                worst = max(samples, key=lambda x: _severity_rank.get(x[1], 0))[1]
                detail_ws.append([
                    bucket_start.isoformat(),
                    service_names.get(sid, str(sid)),
                    round(sum(scores) / len(scores), 2),
                    int(min(scores)),
                    int(max(scores)),
                    len(scores),
                    worst,
                ])
        _finalise(detail_ws)

        meta = wb.create_sheet("Window")
        meta.append(["From", "To", "Bucket Minutes", "Total Snapshots"])
        meta.append([since_aware.isoformat(), until_aware.isoformat(), bucket_minutes, len(snapshots)])
        _finalise(meta)
        return _wb_to_bytes(wb)

    async def service_trend_report(
        self,
        since: datetime,
        until: datetime,
        service_ids: list | None = None,
    ) -> bytes:
        """Per-service score trend: one sheet per service with raw snapshots."""
        since_aware = since if since.tzinfo else since.replace(tzinfo=UTC)
        until_aware = until if until.tzinfo else until.replace(tzinfo=UTC)

        async with self._sf() as session:
            q = (
                select(ServiceScoreSnapshot)
                .where(ServiceScoreSnapshot.captured_at >= since_aware)
                .where(ServiceScoreSnapshot.captured_at <= until_aware)
                .order_by(ServiceScoreSnapshot.captured_at.asc())
            )
            if service_ids:
                q = q.where(ServiceScoreSnapshot.service_id.in_(service_ids))
            snapshots = (await session.execute(q)).scalars().all()
            services = (await session.execute(select(Service))).scalars().all()

        service_names = {s.id: s.name for s in services}
        service_targets = {s.id: s.target_score for s in services}
        per_service: dict = defaultdict(list)
        for snap in snapshots:
            per_service[snap.service_id].append(snap)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # type: ignore[arg-type]

        if not per_service:
            wb.create_sheet("No Data")
            return _wb_to_bytes(wb)

        index_ws = wb.create_sheet("Index")
        index_ws.append([
            "Service", "Target", "Samples", "Avg Score", "Min Score", "Max Score", "Last Score", "Last State",
        ])

        for sid in per_service:
            name = service_names.get(sid, str(sid))
            sheet_name = (name or "service")[:31]
            ws = wb.create_sheet(sheet_name)
            ws.append([
                "Captured At", "Score", "Base Score", "Dependency Penalty", "Health State",
            ])
            scores: list[int] = []
            last = None
            for snap in per_service[sid]:
                ws.append([
                    snap.captured_at.isoformat(),
                    snap.score,
                    snap.base_score,
                    snap.dependency_penalty,
                    snap.health_state,
                ])
                scores.append(snap.score)
                last = snap
            _finalise(ws)
            index_ws.append([
                name,
                service_targets.get(sid),
                len(scores),
                round(sum(scores) / len(scores), 2),
                int(min(scores)),
                int(max(scores)),
                last.score if last else None,
                last.health_state if last else None,
            ])
        _finalise(index_ws)

        meta = wb.create_sheet("Window")
        meta.append(["From", "To", "Total Snapshots", "Services"])
        meta.append([since_aware.isoformat(), until_aware.isoformat(), len(snapshots), len(per_service)])
        _finalise(meta)
        return _wb_to_bytes(wb)


_OPERATOR_FNS = {
    "gt": lambda value, target: value > target,
    "gte": lambda value, target: value >= target,
    "lt": lambda value, target: value < target,
    "lte": lambda value, target: value <= target,
}


def _operator_match(operator: str, value: float, target: float) -> bool:
    fn = _OPERATOR_FNS.get(operator)
    return bool(fn and fn(value, target))


def _pick(values: list[float], fn: ConsolidationFn) -> float | None:
    if not values:
        return None
    if fn == "avg":
        return sum(values) / len(values)
    if fn == "min":
        return min(values)
    if fn == "max":
        return max(values)
    if fn == "p95":
        return percentile(values, 95)
    if fn == "p99":
        return percentile(values, 99)
    if fn == "sum":
        return sum(values)
    if fn == "first":
        return values[0]
    if fn == "last":
        return values[-1]
    return None


def _format_interval(seconds: int) -> str:
    mapping = {
        60: "1 minute",
        300: "5 minutes",
        900: "15 minutes",
        3600: "1 hour",
        21600: "6 hours",
        43200: "12 hours",
        86400: "24 hours",
    }
    return mapping.get(seconds, f"{seconds}s")
